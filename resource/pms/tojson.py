# -*- coding:utf-8 -*-
import os
import json
import datetime

import openpyxl

dirname = os.path.dirname(__file__)

skip_list = ['rep.py']


def get_step():
    wb = openpyxl.load_workbook('dbusid.xlsx')
    ws = wb.active
    max_row = ws.max_row

    info = {}

    for row in range(2, max_row + 1):
        id_ = str(ws.cell(row=row, column=1).value).strip()
        name = str(ws.cell(row=row, column=2).value).strip()
        if id_:
            info[int(id_)] = name

    with open('id2name.json', 'w') as f:
        print(info)
        json.dump(info, f)

    with open('name2id.json', 'w') as f:
        info2 = {value: key for key, value in info.items()}
        json.dump(info2, f)

    info3 = {"baseurl": "http://10.0.10.200:3000/api/v1/zbox/result",
             "product_id": 0,
             "task_id": 0,
             "test_type": "dbus",
             "user_name": "输入自己用户名，如zhangsan"}
    with open('pms_job.json', 'w', encoding='utf8') as f:
        json.dump(info3, f, ensure_ascii=False)

    with open('total.txt', 'a', encoding='utf8') as f:
        content = f"{datetime.datetime.now()}: {len(info)}\n"
        f.write(content)


if __name__ == '__main__':
    get_step()
