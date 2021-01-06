# -*- coding:utf-8 -*-
import os
import time

import openpyxl

dirname = os.path.dirname(__file__)

skip_list = ['rep.py']


def show_file():
    for dir_name, subdir, files in os.walk(dirname):
        files.sort()
        if dir_name == dirname:
            for file in files:
                if file in skip_list:
                    continue
                else:
                    print(os.path.join(dir_name, file))


def rep():
    for dir_name, subdir, files in os.walk(dirname):
        files.sort()
        if '__pycache__' not in dir_name:
            for file in files:
                if file in skip_list:
                    continue
                else:
                    print(os.path.join(dir_name, file))

                isrewrite = False
                with open(os.path.join(dir_name, file), 'r') as f:
                    content_list = f.readlines()
                    for i in range(0, len(content_list)):
                        # frame.base引入
                        if 'frame.base' in content_list[i]:
                            new_content = content_list[i].replace('frame.base', 'core.base_case')
                            if 'OSBase' in new_content:
                                new_content = new_content.replace('OSBase', 'BaseTestCase')
                            content_list[i] = new_content
                            isrewrite = True

                        # 修改 aw.dbus 位置
                        elif 'aw.dbus' in content_list[i]:
                            new_content = content_list[i].replace('aw.dbus', 'apis.dbus_')
                            content_list[i] = new_content
                            isrewrite = True

                        # 修改测试类名
                        elif content_list[i] == 'class testCase(OSBase):\n':
                            content_list[i] = 'class TestCase(BaseTestCase):\n'
                            isrewrite = True

                        # 修改sessionCommon
                        elif 'aw.dbus.sessionBus.sessionCommon' in content_list[i]:
                            new_content = content_list[i].replace('sessionBus.sessionCommon', 'dbus_common')
                            content_list[i] = new_content
                            isrewrite = True

                        # 修改systemCommon
                        elif 'aw.dbus.systemBus.systemCommon' in content_list[i]:
                            new_content = content_list[i].replace('systemBus.systemCommon', 'dbus_common')
                            content_list[i] = new_content
                            isrewrite = True

                if isrewrite:
                    with open(os.path.join(dir_name, file), 'w') as f:
                        for content in content_list:
                            f.write(content)
            else:
                print('+=' * 100)


def get_step(dir_path, m, template_file_path):
    dir_names = dir_path
    # file = os.path.join(os.path.dirname(dirname), 'script/dbus/sessionBus/calendarScheduler/005_updateJob.py')

    step_list = []
    for dir_name, subdir, files in os.walk(dir_names):
        files.sort()
        if '__pycache__' not in dir_name:
            for file in files:
                if file in skip_list:
                    continue
                else:
                    print(os.path.join(dir_name, file))

                step_dict = {"Condition": [], "Step": [], "Result": []}
                flg = 0b0000
                with open(os.path.join(dir_name, file), 'r') as f:
                    content_list = f.readlines()
                    for i in range(0, len(content_list)):
                        if '# @Test Condition:' in content_list[i]:
                            flg = 0b0001
                        if '# @Test Step:' in content_list[i]:
                            flg = 0b0010
                        if '# @Test Result:' in content_list[i]:
                            flg = 0b0100
                        if '# @Test Remark:' in content_list[i]:
                            flg = 0b1000

                        if flg == 0b0001:  # 获取预置条件
                            new_content = content_list[i].replace('# @Test Condition:', '')
                            new_content = new_content.replace('#', '')
                            new_content = new_content.replace(';', '')
                            new_content = new_content.replace('；', '')
                            if new_content.strip():
                                new_content = new_content.strip() + '\n'
                                step_dict["Condition"].append(new_content)

                        if flg == 0b0010:  # 获取步骤
                            new_content = content_list[i].replace('# @Test Step:', '')
                            new_content = new_content.replace('#', '')
                            new_content = new_content.replace(';', '')
                            new_content = new_content.replace('；', '')
                            if new_content.strip():
                                new_content = new_content.strip() + '\n'
                                step_dict["Step"].append(new_content)

                        if flg == 0b0100:  # 获取预期结果
                            new_content = content_list[i].replace('# @Test Result:', '')
                            new_content = new_content.replace('#', '')
                            new_content = new_content.replace(';', '')
                            new_content = new_content.replace('；', '')
                            if new_content.strip():
                                new_content = new_content.strip() + '\n'
                                step_dict["Result"].append(new_content)

                        if flg == 0b1000:  # 处理每项最后一个换行
                            if not step_dict["Condition"]:
                                step_dict["Condition"].append("无\n")
                            for itme in step_dict:
                                len_ = len(step_dict[itme])
                                print(itme)
                                step_dict[itme][len_ - 1] = step_dict[itme][len_ - 1][:-1]

                            break

                    for line in content_list:
                        if '@pytest.mark' not in line:
                            continue

                        if '@pytest.mark.sp2' in line:
                            step_dict["Key"] = 'dbus sp2'

                        elif '@pytest.mark.sp3' in line:
                            step_dict["Key"] = 'dbus sp3'
                        else:
                            step_dict["Key"] = 'dbus'

                        break

                step_dict["Title"] = os.path.splitext(os.path.split(file)[1])[0]
                step_dict["模块"] = m
                # step_dict["Key"] = 'dbus'
                step_dict["优先级"] = 3
                step_dict["用例类型"] = '接口测试'
                step_dict["适用阶段"] = '自动化测试'

                step_list.append(step_dict)
                print(step_dict)

    row = 2
    wb = openpyxl.load_workbook(template_file_path)
    ws = wb["用例库"]

    for num in range(len(step_list)):
        column = 2
        ws.cell(row=row + num, column=column - 1, value=step_list[num]["模块"])
        ws.cell(row=row + num, column=column, value=step_list[num]["Title"])
        ws.cell(row=row + num, column=column + 1, value=''.join(step_list[num]["Condition"]).strip())
        ws.cell(row=row + num, column=column + 2, value=''.join(step_list[num]["Step"]).strip())
        ws.cell(row=row + num, column=column + 3, value=''.join(step_list[num]["Result"]).strip())
        ws.cell(row=row + num, column=column + 4, value=step_list[num]["Key"])
        ws.cell(row=row + num, column=column + 5, value=step_list[num]["优先级"])
        ws.cell(row=row + num, column=column + 6, value=step_list[num]["用例类型"])
        ws.cell(row=row + num, column=column + 7, value=step_list[num]["适用阶段"])

    save_dir = 'xlsx'
    file_name = f'{os.path.split(dir_names)[1]}.xlsx'
    if not os.path.exists('xlsx'):
        os.mkdir(save_dir)
        time.sleep(2)

    wb.save(os.path.join(save_dir, file_name))


def read_template():
    wb = openpyxl.load_workbook('libTemplate.xlsx')
    ws = wb["用例库"]
    cell = ws.cell(row=2, column=4)
    print([cell.value])


if __name__ == '__main__':
    """
    /桌面环境/后端/System Bus/apiLocaleHelper(#25560)
    /桌面环境/后端/System Bus/appsLaunchedRecorder(#25561)
    /桌面环境/后端/System Bus/abRecovery(#25562)
    """
    project_root_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
    dir_path = os.path.join(project_root_path, 'script/dbus/sessionBus/inputDeviceMouse')
    m = '/桌面环境/后端/Session Bus/inputDeviceMouse(#27846)'
    template_file_path = os.path.join('xlsx', 'libTemplate50.xlsx')  # 将pms下载的模板放到xlsx下
    get_step(dir_path, m, template_file_path)
