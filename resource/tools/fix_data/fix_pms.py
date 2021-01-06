# -*- coding: utf-8 -*-
import os
import json

import openpyxl

project_root_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
script_dir = os.path.join(project_root_path, 'script')


def get_case_list():
    global project_root_path
    global script_dir

    case_dict = {}
    for dirname, subdirs, files in os.walk(script_dir):

        files.sort()

        for file in files:
            if file != '__init__.py' and '.py' == os.path.splitext(file)[1]:
                file_path = os.path.join(dirname, file)
                case_dict[file.split('.')[0]] = file_path.split(project_root_path)[1]
    return case_dict


def fix_pms(pms_xlsx):
    case_dict = get_case_list()
    wb = openpyxl.load_workbook(pms_xlsx)
    ws = wb.active
    max_row = ws.max_row

    info = []

    print(f'{"=" * 50}')
    print(f'{"=" * 50}')

    for row in range(2, max_row + 1):
        id_ = str(ws.cell(row=row, column=1).value).strip()
        name = str(ws.cell(row=row, column=2).value).strip()
        if name not in case_dict:
            print(f"{name}")
        else:
            if name in info:
                print(f"发生重复,{name}:{id_}")
            else:
                info.append(name)

    print(f'{"=" * 50}')
    print(f'{"=" * 50}')

    for key in case_dict:
        if key not in info:
            print(f"{key}:{case_dict[key]}")


if __name__ == "__main__":
    fix_pms('./fix_pms/dbusid.xlsx')
